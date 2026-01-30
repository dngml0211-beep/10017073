# -*- coding: utf-8 -*-
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Trying pandas...")
    try:
        import pandas as pd

        file_path = r'c:\우희\02_기획\_2026 기획\북클럽 3.0 기획자료\북클럽3.0_정책 자료\★ 논의사항 정리_260122.xlsx'

        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        print("Available sheets:", excel_file.sheet_names)
        print()

        for sheet_name in excel_file.sheet_names:
            print(f"=== Sheet: {sheet_name} ===")
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            print(df.to_string())
            print("\n")

    except Exception as e:
        print(f"Error with pandas: {e}")
        sys.exit(1)
else:
    file_path = r'c:\우희\02_기획\_2026 기획\북클럽 3.0 기획자료\북클럽3.0_정책 자료\★ 논의사항 정리_260122.xlsx'

    try:
        wb = openpyxl.load_workbook(file_path)
        print("Available sheets:", wb.sheetnames)
        print()

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"=== Sheet: {sheet_name} ===")
            print()

            for row in sheet.iter_rows(values_only=True):
                print('\t'.join([str(cell) if cell is not None else '' for cell in row]))
            print()
            print()
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
