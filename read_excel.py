import openpyxl
import sys

# Open the Excel file
wb = openpyxl.load_workbook(r'c:\우희\02_기획\_2026 기획\북클럽 3.0 기획자료\북클럽3.0_정책 자료\★ 논의사항 정리_260122.xlsx')

# Get all sheet names
print("Available sheets:", wb.sheetnames)
print()

# Read each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print()

    # Print all rows
    for row in sheet.iter_rows(values_only=True):
        print('\t'.join([str(cell) if cell is not None else '' for cell in row]))
    print()
    print()
